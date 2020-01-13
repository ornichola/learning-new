#  https://openpyxl.readthedocs.io/en/stable/index.html

import pprint
import openpyxl as excel

before = excel.open('before.xlsx')
after = excel.open('after.xlsx')

b = before['Sheet1']
a = after['Sheet1']

diff = list()
for i in zip(b.iter_rows(), a.iter_rows()):
    row_before_val = i[0][1].value
    row_after_val = i[1][1].value
    if row_before_val != row_after_val:
        diff.append((i[0][0].value, row_before_val, row_after_val))

# pprint.ppint(diff)

diff_worksheet = excel.Workbook()
diff_sheet = diff_worksheet.active
for i, v in enumerate(diff, start=1):
    diff_sheet[f'A{i}'] = v[0]
    diff_sheet[f'B{i}'] = v[1]
    diff_sheet[f'C{i}'] = v[2]

diff_worksheet.save('diff.xlsx')
